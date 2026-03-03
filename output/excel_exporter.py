"""
excel_exporter.py
-----------------
Exports scored screener results to a formatted Excel workbook.

Sheets:
  1. Grade A      — Top leaders (green header)
  2. Grade B      — Strong stocks (blue header)
  3. Grade C      — Watch list (orange header)
  4. All Results  — Full clean universe
  5. Summary      — Key stats + sector breakdown
"""

import pandas as pd
import os, glob
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RESULTS_DIR = os.path.join(BASE_DIR, "results", "manual")
OUTPUT_DIR  = os.path.join(BASE_DIR, "results", "manual")


# ── Style constants ────────────────────────────────────────────────

GRADE_COLORS = {
    "A": {"header": "0F4C1E", "row_even": "E8F5E9", "row_odd": "F1FAF2"},
    "B": {"header": "1A3A6B", "row_even": "E3EDFA", "row_odd": "EEF4FC"},
    "C": {"header": "7A4000", "row_even": "FFF3E0", "row_odd": "FFF8EE"},
    "D": {"header": "6B1A1A", "row_even": "FDECEA", "row_odd": "FEF3F2"},
}

HEADER_FONT       = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT         = Font(name="Arial", size=10)
BOLD_FONT         = Font(name="Arial", bold=True, size=10)
TITLE_FONT        = Font(name="Arial", bold=True, size=14, color="1A1A2E")
SUBTITLE_FONT     = Font(name="Arial", size=10, color="4B5563")
CENTER_ALIGN      = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN        = Alignment(horizontal="left",   vertical="center")
RIGHT_ALIGN       = Alignment(horizontal="right",  vertical="center")

THIN_BORDER = Border(
    left   = Side(style="thin", color="D1D5DB"),
    right  = Side(style="thin", color="D1D5DB"),
    top    = Side(style="thin", color="D1D5DB"),
    bottom = Side(style="thin", color="D1D5DB"),
)

COLUMNS = [
    ("Symbol",    14, "symbol"),
    ("Grade",      7, "grade"),
    ("RS Rank",   10, "rs_rank"),
    ("RS",        11, "rs"),
    ("Close ₹",   12, "close"),
    ("Tier",      12, "tier"),
    ("Score",      9, "score"),
    ("Sector",    16, "sector"),
    ("RS MA",     11, "rs_ma"),
    ("RS Slope",  11, "rs_slope"),
    ("SMA 20 ₹",  12, "price_sma20"),
    ("WMA 200 ₹", 12, "price_wma200"),
    ("Run Date",  13, "run_date"),
]


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def _write_header_row(ws, grade: str, row: int = 1):
    color  = GRADE_COLORS.get(grade, GRADE_COLORS["A"])["header"]
    fill   = _fill(color)
    for col_idx, (col_name, col_width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = fill
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[row].height = 20


def _write_data_rows(ws, df: pd.DataFrame, grade: str, start_row: int = 2):
    colors = GRADE_COLORS.get(grade, GRADE_COLORS["A"])

    for i, (_, row_data) in enumerate(df.iterrows()):
        row_num = start_row + i
        bg      = colors["row_even"] if i % 2 == 0 else colors["row_odd"]
        fill    = _fill(bg)

        for col_idx, (_, _, field) in enumerate(COLUMNS, start=1):
            val = row_data.get(field, "")

            # Format values
            if field == "rs" and pd.notna(val):
                display = f"+{val:.4f}" if float(val) >= 0 else f"{val:.4f}"
            elif field in ("rs_rank",) and pd.notna(val):
                display = round(float(val), 1)
            elif field in ("rs_ma", "rs_slope") and pd.notna(val):
                display = round(float(val), 4)
            elif field in ("close", "price_sma20", "price_wma200") and pd.notna(val):
                display = round(float(val), 2)
            elif field == "score" and "total_active" in row_data:
                display = f"{int(row_data['score'])}/{int(row_data['total_active'])}"
            else:
                display = val if pd.notna(val) else ""

            cell = ws.cell(row=row_num, column=col_idx, value=display)
            cell.font   = BOLD_FONT if field == "symbol" else BODY_FONT
            cell.fill   = fill
            cell.border = THIN_BORDER

            # Alignment
            if field in ("rs_rank", "rs", "close", "rs_ma",
                         "rs_slope", "price_sma20", "price_wma200"):
                cell.alignment = RIGHT_ALIGN
            elif field in ("grade", "tier", "score"):
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

        ws.row_dimensions[row_num].height = 18

    return start_row + len(df)


def _freeze_and_filter(ws, header_row: int = 1):
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = ws.dimensions


def _add_grade_sheet(wb: Workbook, df: pd.DataFrame, grade: str,
                     sheet_name: str, description: str):
    ws = wb.create_sheet(title=sheet_name)

    # Title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(COLUMNS))
    title_cell = ws.cell(row=1, column=1,
                         value=f"{sheet_name}  —  {description}  ({len(df)} stocks)")
    title_cell.font      = TITLE_FONT
    title_cell.alignment = LEFT_ALIGN
    title_cell.fill      = _fill("F9FAFB")
    ws.row_dimensions[1].height = 28

    # Subtitle
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(COLUMNS))
    sub = ws.cell(row=2, column=1,
                  value=f"Generated: {datetime.today().strftime('%d %b %Y %H:%M')}  |  "
                        f"Sorted by RS Rank (descending)")
    sub.font      = SUBTITLE_FONT
    sub.alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 16

    # Header row at row 3
    _write_header_row(ws, grade, row=3)

    # Data rows from row 4
    _write_data_rows(ws, df, grade, start_row=4)

    # Freeze below header, autofilter
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(COLUMNS))}{3 + len(df)}"

    # Conditional format: RS column (col 4) — red/yellow/green scale
    rs_col = get_column_letter(4)
    rs_range = f"{rs_col}4:{rs_col}{3 + len(df)}"
    ws.conditional_formatting.add(rs_range, ColorScaleRule(
        start_type="min",   start_color="FF4444",
        mid_type="num",     mid_value=0, mid_color="FFFF00",
        end_type="max",     end_color="00AA00",
    ))

    return ws


def _add_summary_sheet(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet(title="Summary", index=0)
    ws.sheet_view.showGridLines = False

    # ── Title ─────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value     = "RS SCREENER — RESULTS SUMMARY"
    t.font      = Font(name="Arial", bold=True, size=16, color="1A1A2E")
    t.alignment = LEFT_ALIGN
    t.fill      = _fill("F0F4FF")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:H2")
    s = ws["A2"]
    s.value     = f"Run date: {datetime.today().strftime('%d %b %Y %H:%M')}   |   Universe: {len(df)} stocks"
    s.font      = SUBTITLE_FONT
    s.alignment = LEFT_ALIGN
    ws.row_dimensions[2].height = 18

    # ── Grade summary ──────────────────────────────────────────────
    ws["A4"] = "GRADE BREAKDOWN"
    ws["A4"].font = Font(name="Arial", bold=True, size=11, color="374151")

    headers = ["Grade", "Count", "Description", "RS Rank Range"]
    grade_info = [
        ("A", "🟢 Top Leaders",    "RS Rank ≥ 80 + Tier 1"),
        ("B", "🔵 Strong",         "RS Rank ≥ 60 or Tier 1"),
        ("C", "🟡 Watch",          "RS Rank ≥ 40 or Tier 2"),
        ("D", "🔴 Avoid",          "Below thresholds"),
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = _fill("374151")
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for i, (grade, desc, rank_desc) in enumerate(grade_info, start=6):
        count = len(df[df["grade"] == grade]) if "grade" in df.columns else 0
        color = GRADE_COLORS.get(grade, {}).get("header", "CCCCCC")
        row_data = [grade, count, desc, rank_desc]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col_idx, value=val)
            cell.font      = BOLD_FONT if col_idx == 1 else BODY_FONT
            cell.fill      = _fill(GRADE_COLORS[grade]["row_even"])
            cell.alignment = CENTER_ALIGN if col_idx <= 2 else LEFT_ALIGN
            cell.border    = THIN_BORDER

    # ── Sector breakdown ───────────────────────────────────────────
    ws["A12"] = "SECTOR BREAKDOWN (Grade A + B)"
    ws["A12"].font = Font(name="Arial", bold=True, size=11, color="374151")

    top_stocks  = df[df["grade"].isin(["A", "B"])] if "grade" in df.columns else df
    sector_dist = top_stocks.groupby("sector").agg(
        Count  = ("symbol", "count"),
        Avg_RS = ("rs", "mean"),
        Tier1  = ("tier", lambda x: (x == "Tier 1").sum()),
    ).sort_values("Count", ascending=False).reset_index()

    sec_headers = ["Sector", "Count", "Avg RS", "Tier 1 Count"]
    for col_idx, h in enumerate(sec_headers, start=1):
        cell = ws.cell(row=13, column=col_idx, value=h)
        cell.font      = HEADER_FONT
        cell.fill      = _fill("374151")
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER

    for i, row_data in sector_dist.iterrows():
        r = 14 + i
        bg = "F9FAFB" if i % 2 == 0 else "FFFFFF"
        vals = [
            row_data["sector"],
            int(row_data["Count"]),
            round(row_data["Avg_RS"], 4),
            int(row_data["Tier1"]),
        ]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font      = BODY_FONT
            cell.fill      = _fill(bg)
            cell.alignment = CENTER_ALIGN if col_idx > 1 else LEFT_ALIGN
            cell.border    = THIN_BORDER

    # ── Column widths ──────────────────────────────────────────────
    for col, width in [("A", 20), ("B", 10), ("C", 28), ("D", 20),
                       ("E", 10), ("F", 10), ("G", 10), ("H", 10)]:
        ws.column_dimensions[col].width = width

    return ws


# ── Main export function ───────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, output_path: str = None) -> str:
    """
    Export scored DataFrame to formatted Excel workbook.

    Parameters
    ----------
    df          : scored DataFrame from scorer.clean_and_score()
    output_path : where to save the file (auto-generated if None)

    Returns
    -------
    str : path to saved Excel file
    """
    if output_path is None:
        ts = datetime.today().strftime("%Y%m%d_%H%M")
        output_path = os.path.join(OUTPUT_DIR, f"rs_results_{ts}.xlsx")

    # Ensure is_clean exists
    if "is_clean" not in df.columns:
        df["is_clean"] = True

    clean_df = df[df["is_clean"]].copy()

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── Summary sheet (first) ──────────────────────────────────────
    _add_summary_sheet(wb, clean_df)

    # ── Grade sheets ───────────────────────────────────────────────
    grade_configs = [
        ("A", "Grade A — Top Leaders",  "🟢 RS Rank ≥ 80, Tier 1"),
        ("B", "Grade B — Strong",       "🔵 RS Rank ≥ 60 or Tier 1"),
        ("C", "Grade C — Watch",        "🟡 RS Rank ≥ 40 or Tier 2"),
    ]

    for grade, sheet_name, description in grade_configs:
        grade_df = clean_df[clean_df["grade"] == grade].copy()
        if not grade_df.empty:
            _add_grade_sheet(wb, grade_df, grade, sheet_name, description)

    # ── All clean results ──────────────────────────────────────────
    ws_all = wb.create_sheet(title="All Results")
    ws_all.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(COLUMNS))
    t = ws_all.cell(row=1, column=1,
                    value=f"All Clean Results  —  {len(clean_df)} stocks")
    t.font      = TITLE_FONT
    t.alignment = LEFT_ALIGN
    t.fill      = _fill("F9FAFB")
    ws_all.row_dimensions[1].height = 28

    _write_header_row(ws_all, "A", row=2)
    _write_data_rows(ws_all, clean_df, "A", start_row=3)
    ws_all.freeze_panes = "A3"
    ws_all.auto_filter.ref = (
        f"A2:{get_column_letter(len(COLUMNS))}{2 + len(clean_df)}"
    )

    wb.save(output_path)
    return output_path


# ── Load latest scored CSV ─────────────────────────────────────────

def load_latest_scored() -> pd.DataFrame | None:
    files = sorted(
        glob.glob(os.path.join(RESULTS_DIR, "scored_*.csv")),
        reverse=True
    )
    if not files:
        # Fall back to raw screener CSV
        files = sorted(
            glob.glob(os.path.join(RESULTS_DIR, "screener_*.csv")),
            reverse=True
        )
    if not files:
        return None, None
    path = files[0]
    return pd.read_csv(path), path


# ── Verification test ──────────────────────────────────────────────

if __name__ == "__main__":

    print("=" * 60)
    print("  EXCEL EXPORTER — VERIFICATION TEST")
    print("=" * 60)

    df, source_path = load_latest_scored()

    if df is None:
        print("  No scored CSV found. Run screener.py + scorer.py first.")
        exit()

    print(f"  Source : {os.path.basename(source_path)}")
    print(f"  Rows   : {len(df)}")

    # Score if needed
    if "grade" not in df.columns:
        print("  Scoring results first...")
        import sys
        sys.path.insert(0, BASE_DIR)
        from core.scorer import clean_and_score
        df = clean_and_score(df)

    print("  Exporting to Excel...")
    out = export_to_excel(df)

    print(f"\n  ✅ Saved: {out}")
    print(f"  Sheets: Summary | Grade A | Grade B | Grade C | All Results")
    print(f"  Open in Excel or Numbers to view.")
